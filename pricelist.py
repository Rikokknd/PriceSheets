from copy import deepcopy
import random
from singleton_decorator import singleton
from datetime import datetime, timedelta
import time
import logging
import sys
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import gspread
import requests
import io
import os
from gspread_formatting import *

URL = "https://89.248.193.157:65002/price/PRC%20(XLSX).xlsx"
FOLDER = "/var/www/u0853380/data/priceSheets/"
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)


def grab_url_file_to_memory(url):
    """Берем файл по ссылке и записываем в память, возвращаем в виде потока байтов."""
    r = requests.get(url, verify=False)
    f = io.BytesIO(r.content)
    return f


def xlsx_compare(first, second):
    """Принимаем 2 xlsx-файла, снимаем объединения ячеек в обоих.
    Перегоняем данные из ячеек в массивы. Возвращаем результат сравнения массивов."""
    wb_first = openpyxl.load_workbook(first).active

    for merge in list(wb_first.merged_cells):
        wb_first.unmerge_cells(range_string=str(merge))

    wb_second = openpyxl.load_workbook(second).active

    for merge in list(wb_second.merged_cells):
        wb_second.unmerge_cells(range_string=str(merge))

    row_new_list = []
    row_old_list = []

    for row_new, row_old in zip(wb_first.iter_rows(), wb_second.iter_rows()):
        [row_new_list.append([cell_new.coordinate, cell_new.internal_value]) for cell_new in row_new]
        [row_old_list.append([cell_old.coordinate, cell_old.internal_value]) for cell_old in row_old]

    return row_new_list == row_old_list


@singleton
def logger():
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s',
                        level=logging.INFO, filename=FOLDER + f'logs/{datetime.now().strftime("%m_%d_%Y")}.log',
                        filemode='a')
    log_maker = logging.getLogger(__name__)
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    log_maker.addHandler(handler)
    return log_maker


@singleton
class Header:
    """Содержит информацию для заголовка страниц, дату последнего обновления, наименования столбцов."""

    def __init__(self):
        self._header_text_rows = {'phone': "Телефон: 071-312-3-777, Эдгар",
                                  'address': "ул. Буденновских партизан, 83а(рынок Объединенный, кольцо трамвая)",
                                  'hours': "Ежедневно 10.00-18.00",
                                  'DT': f"Последнее обновление: {datetime.now().strftime('%H:%M %d/%m')}"}
        self._col_dict = {}
        self._keywords = []
        self._set_keywords()

    def parse_header(self, *rows):
        """Получаем несколько рядов из заголовка, считываем названия колонок, возвращаем объект Header."""
        for row in rows:
            for cell in row:
                if cell.value and cell.value != 'Цена':
                    self._col_dict[cell.column] = cell.value
        return self

    def _set_keywords(self):
        self._keywords = [["Xiaomi", "Redmi"], ["iPhone", "iPh"], ["Huawei"], ["Samsung"],
                          ["Realme", "Oppo"], ["Meizu"], ['Nokia'], ['ZTE'], ['SONY'],
                          ['LENOVO'], ['onePlus'], ['LeEco'], ['Разборка телефонов', 'Б/У'], ["SSD"], ["Остальное"]]

    def get_keywords(self):
        return self._keywords

    def get_header_text(self):
        return self._header_text_rows

    def get_col_headers(self):
        return self._col_dict

    def get_col_headers_cleaned(self) -> list:
        clean = []
        description = False
        for key, value in self._col_dict.items():
            if value == 'Описание':
                description = True
                continue
            if value is not None:
                clean.append(value)
        if description:
            clean.append('Описание')
        return clean


@singleton
class GoogleSpreadsheetEditor:
    def __init__(self, update_only_time=False):
        self.spreadsheet = self.auth_workbook()
        self.sheet_titles = ['\\'.join(x) for x in Header().get_keywords()]
        self._check_worksheets()
        if update_only_time:
            self.update_time()

    @staticmethod
    def auth_workbook():
        scope = ["https://spreadsheets.google.com/feeds",
                 "https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file",
                 "https://www.googleapis.com/auth/drive",
                 "https://www.googleapis.com/auth/drive.readonly",
                 "https://www.googleapis.com/auth/spreadsheets.readonly"]
        credentials = ServiceAccountCredentials.from_json_keyfile_name(FOLDER + "creds.json", scope)
        client = gspread.authorize(credentials)
        book = client.open("Запчасти для телефонов, ноутбуков")
        return book

    def _check_worksheets(self):
        """Проверяем, чтобы все страницы были на местах. При необходимости удаляем/добавляем/переименовываем."""
        sheets = self.spreadsheet.worksheets()
        # если страниц слишком мало - добавляем
        if len(sheets) < len(self.sheet_titles):
            for i in range(len(self.sheet_titles) - len(sheets)):
                self.spreadsheet.add_worksheet(str(random.randint(1, 999999)), 1000, 10)
        # если страниц слишком много - удаляем
        elif len(sheets) > len(self.sheet_titles):
            for i in range(len(self.sheet_titles), len(sheets)):
                self.spreadsheet.del_worksheet(sheets[i])

        sheets = self.spreadsheet.worksheets()
        for i in range(len(self.sheet_titles)):
            if self.sheet_titles[i] in [sheet.title for sheet in sheets]:
                self.spreadsheet.worksheet(self.sheet_titles[i]).update_index(i)
            else:
                sheets[i].update_title(self.sheet_titles[i])

    def update_sheet(self, sheet_index, header_dict, update, formats):
        sheet = self.spreadsheet.get_worksheet(sheet_index)
        sheet.delete_columns(1, 8)
        sheet.add_cols(0)
        sheet.freeze(rows=3)
        sheet.batch_update(header_dict['batch_update'])
        set_column_widths(sheet, header_dict['set_column_widths'])
        set_row_heights(sheet, header_dict['set_row_heights'])
        format_cell_ranges(sheet, header_dict['format_cell_ranges'])
        for el in header_dict['merge_cells_iterate']:
            sheet.merge_cells(el)
        sheet.batch_update(update)
        if len(formats) > 0:
            format_cell_ranges(sheet, formats)

    def update_time(self):
        """Обновляем только время последнего обновления в ячейке B1 на каждой странице"""
        dt = f"Последнее обновление: {datetime.now().strftime('%H:%M %d/%m')}"
        for sheet_index in range(len(self.sheet_titles)):
            sheet = self.spreadsheet.get_worksheet(sheet_index)
            sheet.batch_update([{
                'range' : 'B1',
                'values' : [[dt]]
            }])
            time.sleep(2)


class Item:
    """Товар из прайса. Содержит название, остаток и цены на товар, принадлежность к какой-либо странице прайса."""

    def __init__(self, item_row):
        self._properties = {'page': []}
        self._make_item(item_row)
        self.get_item()

    def _make_item(self, item_row):
        cols = Header().get_col_headers()  # тянем список названий колонок
        for i in range(len(item_row)):  # бежим по ячейкам ряда
            if item_row[i].value is None:  # если ячейка пустая, пропуск
                continue
            self._properties[cols[i + 1]] = item_row[i].value  # привязываем значение ячейки к названию столбца
        keywords = Header().get_keywords()  # тянем список названий страниц
        # Если в названии товара содержится ключевое слово из названия страницы - добавляем принадлежность этой странице
        for i in range(len(keywords)):
            if any(brand_name.lower() in self._properties['Номенклатура'].lower() for brand_name in keywords[i]):
                self._properties['page'].append(i)
        if len(self._properties['page']) == 0:
            # если ни одного слова не совпало - добавляем на последнюю страницу
            self._properties['page'].append(len(keywords) - 1)

    def get_item(self):
        return self

    def get_item_props(self) -> dict:
        """возвращает dict вида: {
        'Наименование' : 'что-то',
        'Остаток' : int,
        'Какой-то вид цены' : int,
        'page' : list[int*]
        }"""
        return self._properties


class ItemGroup(object):
    """Группа, в которой содержится папка из прайса. Может содержать под-группы и товары.
    Содержит имя группы и ссылки на родительский и дочерние эл-ты."""

    def __init__(self):
        self.parent = None
        self.header_row = None
        self.children_list = []

    def __iter__(self):
        return iter(self.children_list)

    def __len__(self):
        return len(self.children_list)

    def remove_element(self, element):
        self.children_list = [el for el in self.children_list if el is not element]

    def set_parent(self, new_parent):
        self.parent = new_parent

    def set_header(self, header_row):
        self.header_row = header_row[0].value

    def get_header(self):
        return self.header_row

    def add_child(self, child):
        if type(child) == tuple:
            # создаем экземпляр Item
            self.children_list.append(Item(child))
        else:
            self.children_list.append(child)

    def move_children_to_parent(self):
        """Передаем все дочерние эл-ты родителю, обнуляем список дочерних эл-тов"""
        for child in self.children_list:
            self.parent.add_child(child)
        self.children_list = []

    def raiser(self):
        """Проверяет входящие группы, и если в этих группах всего 1 элемент,
        который тоже является группой - забираем из него детей и перекладываем в группу выше"""
        elements_to_remove = []
        # бежим по дочерним элементам группы
        for element in self.children_list:
            if type(element) == Item:
                continue
            if len(element) == 0:
                elements_to_remove.append(element)
            if len(element) == 1:
                # если содержит только 1 child и это группа - запускаем в ней raiser
                if type(element.children_list[0]) == ItemGroup:
                    element.children_list[0].raiser()
                    # запускаем передачу эл-тов в родительскую группу
                    element.children_list[0].move_children_to_parent()
                    elements_to_remove.append(element.children_list[0])
            for el in elements_to_remove:
                element.remove_element(el)


class PriceListPage:
    """Содержит заголовок страницы и группы с товарами, которые попадают на эту страницу.
    Принимает КОПИЮ групп. После получения просеивает группы, оставляя только те что должны попасть на страницу."""

    def __init__(self, page_index: int, page_name: list, groups: ItemGroup):
        self._index = page_index
        self.name = "\\".join(page_name)
        self._groups = groups
        self.sift_groups()

    def sift_groups(self):
        def sieve(group: ItemGroup):
            """Сито, которое отсеивает неподходящие элементы и пустые группы"""
            elements_to_remove = []
            for item in group:
                if type(item) == Item:
                    # если Item не принадлежит этой странице, добавляем в список на удаление
                    if self._index not in item.get_item_props()['page']:
                        elements_to_remove.append(item)
                # Если элемент - вложенная группа, ее тоже загоняем в сито.
                elif type(item) == ItemGroup:
                    # добавляем в список на удаление если группа возвращается, если не удаляется то возвращается None
                    elements_to_remove.append(sieve(item))
                else:
                    raise TypeError(f"Member of group cannot be {type(item)}!")
            for el in elements_to_remove:
                if el is not None:
                    group.remove_element(el)
            # возвращаем группу, если она опустела.
            if len(group) == 0:
                return group
            else:
                return None

        sieve(self._groups)
        self._groups.raiser()

    def get_content(self):
        return self._groups


@singleton
class PriceList:
    """Корневой класс. Содержит заголовок, список ключевых слов для генерации страниц прайса,
    список групп, страницы прайса, позже добавлю еще что-нибудь."""

    def __init__(self, log_machine, link_to_file=None):
        self._logger = log_machine
        self._worksheet = openpyxl.load_workbook(link_to_file).active
        self._header = Header().parse_header(self._worksheet[1], self._worksheet[2])
        self._groups = self._parse_groups()
        self._sheet_keywords = Header().get_keywords()
        self._item_pages = self._create_pages()
        self._editor = GoogleSpreadsheetEditor()
        self.send_pages()
        self._logger.info("Finished.")

    def _parse_groups(self) -> ItemGroup:
        """Запускает поиск групп в документе. Возвращает список групп."""

        def cleanup(x: list):
            """Обрабатывает результат работы splitter.
            Превращает все tuple в многомерном списке в одиночные элементы с номером ряда"""
            result = []
            for i in x:
                if type(i) is list:
                    result.append(cleanup(i))
                else:
                    result.append(i[1])
            return result

        def splitter(x: list, base=0):
            """Обрабатывает список, состоящий из tuple(уровень ряда, номер ряда).
            Возвращает многоуровневый список групп вида [название группы, [элемент группы, элемент группы...]]"""
            result = []
            if x[0][0] < base:
                result.append(x[0])
            if len(set([a[0] for a in x[1:]])) == 1:
                result.append(x[1:])
                return result
            for i in range(len(x)):
                if x[i][0] == base:
                    for y in range(i + 1, len(x)):
                        if x[y][0] == base:
                            result.append(splitter(x[i:y], base + 1))
                            break
                    else:
                        result.append(splitter(x[i:], base + 1))
            return result

        # собираем список рядов в виде tuple(уровень_ряда, номер_ряда)
        self._logger.info("Pulling groups and items from price list...")
        outline_levels = []
        for row in range(3, self._worksheet.max_row + 1):
            outline_levels.append((self._worksheet.row_dimensions[row].outline_level, row))
        groups = cleanup(splitter(outline_levels))
        groups.insert(0, 1)
        return self._group_maker(groups)

    def _group_maker(self, row_list: list, parent=None) -> ItemGroup:
        """Конструктор групп, умеет в рекурсию.
        Обрабатывает список вида [название группы, [элемент группы, элемент группы...]]"""
        new_group = ItemGroup()
        if parent is not None:
            new_group.set_parent(parent)
        for row in range(len(row_list)):
            if row == 0:
                new_group.set_header(self._worksheet[row_list[row]])
                continue
            # если внутри списка элементов есть еще хоть один список - значит этот список тоже группа
            if any(type(x) == list for x in row_list[row]):
                # вторым аргументом отправляется родительский элемент
                new_group.add_child(self._group_maker(row_list[row], new_group))
            else:
                for item in row_list[row]:
                    new_group.add_child(self._worksheet[item])  # конструктор add_child возвращает Item
        return new_group

    def _create_pages(self):
        self._logger.info("Distributing groups and items to sheets...")
        item_pages = []
        for i in range(len(self._sheet_keywords)):
            item_pages.append(PriceListPage(i, self._sheet_keywords[i], deepcopy(self._groups)))
        return item_pages

    def send_pages(self):
        def compose_items(group: ItemGroup, cols: list) -> list:
            result = []
            for item in group:
                if type(item) == ItemGroup:
                    result.append([item.get_header()])
                    result += compose_items(item, cols)
                elif type(item) == Item:
                    row = []
                    props = item.get_item_props()
                    for col_name in cols:
                        if col_name in props:
                            row.append(props[col_name])
                        else:
                            row.append(None)
                    result.append(row)
            return result

        # секция, в которой генерируется шапка, одинаковая на каждой странице
        # В словаре header_dict содержатся: содержимое ячеек(batch_update),
        # ширина и высота столбцов/строк(set_column_widths, set_row_heights),
        # параметры стилей ячеек для шапки(format_cell_ranges), ячейки которые нужно объединить(merge_cells_iterate)
        header_dict = {}
        header_text = self._header.get_header_text()
        col_names = self._header.get_col_headers_cleaned()
        header_dict['batch_update'] = [
            {'range': 'A1:B1', 'values': [[header_text['phone'], header_text['DT']]]},
            {'range': 'A2:B2', 'values': [[header_text['address'], header_text['hours']]]},
            {'range': f'A3:{"ABCDEFGH"[len(col_names) - 1]}3', 'values': [[x for x in col_names]]}]
        header_dict['set_column_widths'] = [('A', 1000), ("B:H", 100), (f"{'ABCDEFGH'[len(col_names) - 1]}", 300)]
        header_dict['set_row_heights'] = [('1:2', 25), ('3', 50), ('4:500', 25)]
        header_dict['format_cell_ranges'] = [
            (f'B4:B500', CellFormat(horizontalAlignment='CENTER',
                                    verticalAlignment='MIDDLE',
                                    wrapStrategy='Wrap')),
            (f'A1:{"ABCDEFGH"[len(col_names) - 1]}1', CellFormat(backgroundColor=Color(0.91, 0.26, 0.21),  # orange red
                                                                 textFormat=TextFormat(bold=True,
                                                                                       foregroundColor=Color(0, 0, 0),
                                                                                       fontFamily="Roboto",
                                                                                       fontSize=11),
                                                                 horizontalAlignment='CENTER',
                                                                 verticalAlignment='MIDDLE',
                                                                 wrapStrategy='Wrap')),
            (f'A2:{"ABCDEFGH"[len(col_names) - 1]}2', CellFormat(backgroundColor=Color(1, 0.84, 0.4),  # soft yellow
                                                                 textFormat=TextFormat(bold=True,
                                                                                       foregroundColor=Color(0, 0, 0),
                                                                                       fontFamily="Roboto",
                                                                                       fontSize=11),
                                                                 horizontalAlignment='CENTER',
                                                                 verticalAlignment='MIDDLE',
                                                                 wrapStrategy='Wrap')),
            (f'A3:{"ABCDEFGH"[len(col_names) - 1]}3', CellFormat(backgroundColor=Color(0, 0, 1),  # blue
                                                                 textFormat=TextFormat(bold=True,
                                                                                       foregroundColor=Color(0, 0, 0),
                                                                                       fontFamily="Arial", fontSize=12),
                                                                 horizontalAlignment='CENTER',
                                                                 verticalAlignment='MIDDLE',
                                                                 wrapStrategy='Wrap'))]
        header_dict['merge_cells_iterate'] = [f"B1:{'ABCDEFGH'[len(col_names) - 1]}1",
                                              f"B2:{'ABCDEFGH'[len(col_names) - 1]}2"]
        # конец секции с шапкой

        group_row_style = CellFormat(backgroundColor=Color(1, 0.43, 0),  # dark orange
                                     textFormat=TextFormat(bold=True),
                                     horizontalAlignment='LEFT', verticalAlignment='MIDDLE')
        # item_row_style = CellFormat(backgroundColor=Color(1, 1, 0),  # yellow
        #                             textFormat=TextFormat(bold=False),
        #                             horizontalAlignment='LEFT', verticalAlignment='MIDDLE')
        self._logger.info("Sending sheets to project with 10s timeout to prevent API overload.")
        for i in range(len(self._item_pages)):  # цикл по страницам
            batch_update = []
            cell_formats = []
            rows = compose_items(self._item_pages[i].get_content(), col_names)
            for y in range(len(rows)):  # цикл по рядам
                if len(rows[y]) == 1:  # если 1 ячейка - это заголовок группы
                    batch_update.append({'range': f'A{len(header_dict["batch_update"]) + y + 1}',
                                         'values': [[rows[y][0]]]})
                    cell_formats.append((
                        f'A{len(header_dict["batch_update"]) + y + 1}:{"ABCDEFGH"[len(col_names) - 1]}{len(header_dict["batch_update"]) + y + 1}',
                        group_row_style,))
                else:  # иначе это товар
                    batch_update.append({
                                            'range': f'A{len(header_dict["batch_update"]) + y + 1}:{"ABCDEFGH"[len(col_names) - 1]}{len(header_dict["batch_update"]) + y + 1}',
                                            'values': [rows[y]]})
                    # cell_formats.append((
                    #                     f'A{len(header_dict["batch_update"]) + y + 1}:{"ABCDEFGH"[len(col_names) - 1]}{len(header_dict["batch_update"]) + y + 1}',
                    #                     item_row_style,))
            self._editor.update_sheet(i, header_dict, batch_update, cell_formats)
            self._logger.info(f"Sheet {self._item_pages[i].name} sent.")
            if i + 1 != len(self._item_pages):
                time.sleep(10)


if __name__ == '__main__':
    log = logger()
    log.info("============================================================")
    log.info("Obtaining file...")
    new_file_data = grab_url_file_to_memory(URL)
    log.info("Got it.")
    if os.path.isfile(FOLDER + "pricelist.xlsx"):  # проверить, существует ли файл pricelist.xlsx
        saved_file_data = io.BytesIO(open(FOLDER + "pricelist.xlsx", 'rb').read())  # читаем существующий файл в память
        log.info("Comparing new file to old...")
        if xlsx_compare(new_file_data, saved_file_data):
            GoogleSpreadsheetEditor(True)
            log.info("No changes. Time updated. Shutting down.")
            sys.exit()
        else:
            log.info("Changes found, updating.")
            PriceList(log, new_file_data)
            log.info("Update finished. Saving file...")
            with open(FOLDER + "pricelist.xlsx", 'wb') as outfile:
                outfile.write(new_file_data.getvalue())
            log.info("Saved. Shutting down.")
    else:
        log.info("Updating...")
        PriceList(log, new_file_data)
        log.info("Update finished. Saving file...")
        with open(FOLDER + "pricelist.xlsx", 'wb') as outfile:
            outfile.write(new_file_data.getvalue())
        log.info("Saved. Shutting down.")
